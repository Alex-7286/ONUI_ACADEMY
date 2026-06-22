import json
import re
from pathlib import Path
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from openpyxl import load_workbook
from pydantic import BaseModel

from app.core.database import get_conn
from app.core.jwt_token import decode_access_token

router = APIRouter(prefix="/exams", tags=["exams"])
security = HTTPBearer(auto_error=False)

BACKEND_DIR = Path(__file__).resolve().parents[3]
LECTURE_DIR = BACKEND_DIR / "data" / "lecture"

EXAM_SHEET_NAMES = {
    "midterm": "중간고사",
    "final": "기말고사",
}

OPTION_INDEX = {
    "①": 0,
    "②": 1,
    "③": 2,
    "④": 3,
}

OPTION_RE = re.compile(r"^([①②③④])\s*(.*)$")


class ExamAnswer(BaseModel):
    question_id: str
    selected_index: int | None = None


class ExamSubmission(BaseModel):
    level: str = "초급 1"
    exam_type: str = "midterm"
    answers: list[ExamAnswer]


def _optional_user_id(credentials: HTTPAuthorizationCredentials | None) -> int | None:
    if not credentials or credentials.scheme.lower() != "bearer":
        return None

    try:
        payload = decode_access_token(credentials.credentials)
        return int(payload.get("sub", "0"))
    except (ValueError, TypeError):
        return None


def _normalize(value: str) -> str:
    return re.sub(r"\s+", "", value).lower()


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _find_excel_path(level: str) -> Path | None:
    normalized_level = _normalize(level)
    candidates = [path for path in LECTURE_DIR.glob("*.xlsx") if not path.name.startswith("~$")]

    for path in candidates:
        if normalized_level in _normalize(path.stem):
            return path

    compact_level = normalized_level.replace(" ", "")
    for path in candidates:
        if compact_level in _normalize(path.stem):
            return path

    return None


def _parse_question(raw_question: str) -> tuple[str, str, list[str]]:
    lines = [line.strip() for line in raw_question.replace("\r\n", "\n").splitlines() if line.strip()]
    if not lines:
        return "", "", []

    prompt = re.sub(r"^Q\.\s*", "", lines[0]).strip()
    passage_lines: list[str] = []
    options: list[str] = []

    for line in lines[1:]:
        option_match = OPTION_RE.match(line)
        if option_match:
            options.append(option_match.group(2).strip())
            continue

        if options:
            options[-1] = f"{options[-1]}\n{line}".strip()
        else:
            passage_lines.append(line.removeprefix("▶").strip())

    return prompt, "\n".join(passage_lines).strip(), options


def _load_exam_questions(level: str, exam_type: str) -> list[dict[str, Any]]:
    excel_path = _find_excel_path(level)
    if excel_path is None:
        raise HTTPException(status_code=404, detail="해당 레벨의 시험 엑셀을 찾을 수 없습니다.")

    sheet_keyword = EXAM_SHEET_NAMES.get(exam_type)
    if sheet_keyword is None:
        raise HTTPException(status_code=400, detail="지원하지 않는 시험 유형입니다.")

    workbook = load_workbook(excel_path, read_only=True, data_only=True)

    try:
        worksheet = next((sheet for sheet in workbook.worksheets if sheet_keyword in sheet.title), None)
        if worksheet is None:
            raise HTTPException(status_code=404, detail="시험 시트를 찾을 수 없습니다.")

        questions: list[dict[str, Any]] = []

        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, max_col=6, values_only=True), start=2):
            week, lesson, _, question_type, raw_question, raw_answer = row
            question_text = _cell_text(raw_question)
            answer_text = _cell_text(raw_answer)
            if not question_text:
                continue

            prompt, passage, options = _parse_question(question_text)
            if not prompt or len(options) < 2:
                continue

            answer_symbol = next((symbol for symbol in OPTION_INDEX if symbol in answer_text), "")
            correct_index = OPTION_INDEX.get(answer_symbol)
            if correct_index is None:
                continue

            questions.append(
                {
                    "id": f"{exam_type}-{row_number}",
                    "week": int(week) if week is not None else None,
                    "lesson": int(lesson) if lesson is not None else None,
                    "type": _cell_text(question_type),
                    "prompt": prompt,
                    "passage": passage,
                    "options": options,
                    "correctIndex": correct_index,
                }
            )

        return questions
    finally:
        workbook.close()


@router.get("/questions")
def list_exam_questions(
    level: str = "초급 1",
    exam_type: str = Query(default="midterm", alias="type"),
) -> dict[str, Any]:
    questions = _load_exam_questions(level, exam_type)

    return {
        "level": level,
        "type": exam_type,
        "title": EXAM_SHEET_NAMES[exam_type],
        "durationMinutes": 60,
        "questionCount": len(questions),
        "questions": [
            {
                key: value
                for key, value in question.items()
                if key != "correctIndex"
            }
            for question in questions
        ],
    }


@router.get("/status")
def get_exam_status(
    level: str = "초급 1",
    exam_type: str = Query("midterm", alias="type"),
    credentials: HTTPAuthorizationCredentials | None = Depends(security),
) -> dict[str, Any]:
    student_id = _optional_user_id(credentials)
    if student_id is None:
        return {"submitted": False, "canTake": True}

    with get_conn() as conn:
        row = conn.execute(
            """
            SELECT id, score, submitted_at
            FROM exam_submissions
            WHERE student_id = ? AND level = ? AND exam_type = ?
            ORDER BY id ASC
            LIMIT 1
            """,
            (student_id, level, exam_type),
        ).fetchone()

    return {
        "submitted": bool(row),
        "canTake": not bool(row),
        "score": row["score"] if row else None,
        "submittedAt": row["submitted_at"] if row else None,
    }


@router.post("/submit")
def submit_exam(
    submission: ExamSubmission,
    request: Request,
    credentials: HTTPAuthorizationCredentials | None = Depends(security),
) -> dict[str, Any]:
    questions = _load_exam_questions(submission.level, submission.exam_type)
    answer_map = {answer.question_id: answer.selected_index for answer in submission.answers}
    correct_count = sum(
        1
        for question in questions
        if answer_map.get(question["id"]) == question["correctIndex"]
    )
    question_count = len(questions)
    answered_count = sum(answer.selected_index is not None for answer in submission.answers)
    score = round((correct_count / question_count) * 100) if question_count else 0
    student_id = _optional_user_id(credentials)

    with get_conn() as conn:
        if student_id is not None:
            existing = conn.execute(
                """
                SELECT id
                FROM exam_submissions
                WHERE student_id = ? AND level = ? AND exam_type = ?
                ORDER BY id ASC
                LIMIT 1
                """,
                (student_id, submission.level, submission.exam_type),
            ).fetchone()
            if existing:
                exam_name = EXAM_SHEET_NAMES.get(submission.exam_type, "시험")
                raise HTTPException(
                    status_code=409,
                    detail=f"{submission.level} {exam_name}는 이미 응시했습니다. 중간고사와 기말고사는 한 번만 응시할 수 있습니다.",
                )

        cur = conn.execute(
            """
            INSERT INTO exam_submissions (
                student_id,
                level,
                exam_type,
                question_count,
                answered_count,
                correct_count,
                score,
                answers_json,
                ip_address
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                student_id,
                submission.level,
                submission.exam_type,
                question_count,
                answered_count,
                correct_count,
                score,
                json.dumps([answer.model_dump() for answer in submission.answers], ensure_ascii=False),
                request.client.host if request.client else None,
            ),
        )
        conn.commit()

    return {
        "submissionId": cur.lastrowid,
        "questionCount": question_count,
        "answeredCount": answered_count,
        "correctCount": correct_count,
        "score": score,
        "passed": score >= 70,
    }
