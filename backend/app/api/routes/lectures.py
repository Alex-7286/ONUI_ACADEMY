import json
import re
from pathlib import Path
from typing import Any

from fastapi import APIRouter
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

router = APIRouter(prefix="/lectures", tags=["lectures"])

BACKEND_DIR = Path(__file__).resolve().parents[3]
LECTURE_DIR = BACKEND_DIR / "data" / "lecture"
CACHE_DIR = BACKEND_DIR / "data" / "cache" / "lecture_slides"
LECTURE_IMAGE_DIR = BACKEND_DIR.parent / "frontend" / "public" / "images"
WEEK_TEXT = "주차"
LESSON_TEXT = "차시"
BEGINNER_1_TEXT = "초급 1"
BEGINNER_1_COMPACT_TEXT = "초급1"
LESSON_TEXT_MARKER = "교안"
SCRIPT_TEXT_MARKER = "스크립트"
VOCABULARY_TEXT_MARKER = "단어장"
SUMMARY_TEXT_MARKER = "강의요약본"
IMAGE_TEXT_MARKER = "image"
LESSON_HEADER_RE = re.compile(r"(\d+)\s*" + WEEK_TEXT + r"\s*(\d+)\s*" + LESSON_TEXT + r"\s*-\s*(.+)")
SUMMARY_HEADER_RE = re.compile(r"(\d+)\s*" + WEEK_TEXT + r"\s*(\d+)\s*" + LESSON_TEXT)
IMAGE_RE = re.compile(r"^[^\s]+\.(?:png|jpe?g|webp|gif|svg)$", re.IGNORECASE)

CACHE_VERSION = 6

_CACHE: dict[str, tuple[int, dict[str, Any]]] = {}


def _normalize(value: str) -> str:
    return re.sub(r"\s+", "", value).lower()


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _number(value: Any, fallback: int) -> int:
    text = _cell_text(value)
    match = re.search(r"\d+", text)
    return int(match.group(0)) if match else fallback


def _excel_candidates() -> list[Path]:
    return [path for path in LECTURE_DIR.glob("*.xlsx") if not path.name.startswith("~$")]


def _find_excel_path(level: str) -> Path | None:
    normalized_level = _normalize(level)
    candidates = _excel_candidates()

    for path in candidates:
        if normalized_level in _normalize(path.stem):
            return path

    if normalized_level == _normalize(BEGINNER_1_TEXT):
        for path in candidates:
            if _normalize(BEGINNER_1_COMPACT_TEXT) in _normalize(path.stem):
                return path

    return None


def _find_lesson_sheet(workbook: Any, level: str) -> Worksheet | None:
    normalized_level = _normalize(level)

    for worksheet in workbook.worksheets:
        normalized_title = _normalize(worksheet.title)
        if normalized_level in normalized_title and LESSON_TEXT_MARKER in worksheet.title and SCRIPT_TEXT_MARKER in worksheet.title:
            return worksheet

    for worksheet in workbook.worksheets:
        if LESSON_TEXT_MARKER in worksheet.title and SCRIPT_TEXT_MARKER in worksheet.title:
            return worksheet

    return None


def _find_summary_sheet(workbook: Any, level: str) -> Worksheet | None:
    normalized_level = _normalize(level)
    normalized_marker = _normalize(SUMMARY_TEXT_MARKER)

    for worksheet in workbook.worksheets:
        normalized_title = _normalize(worksheet.title)
        if normalized_level in normalized_title and normalized_marker in normalized_title:
            return worksheet

    for worksheet in workbook.worksheets:
        if normalized_marker in _normalize(worksheet.title):
            return worksheet

    return None


def _lesson_image_index(level: str) -> dict[str, str]:
    compact_level = re.sub(r"\s+", "", level)
    level_directory = LECTURE_IMAGE_DIR / compact_level

    if not level_directory.is_dir() and LECTURE_IMAGE_DIR.is_dir():
        normalized_level = _normalize(level)
        level_directory = next(
            (
                directory
                for directory in LECTURE_IMAGE_DIR.iterdir()
                if directory.is_dir() and _normalize(directory.name) in normalized_level
            ),
            level_directory,
        )

    if not level_directory.is_dir():
        return {}

    return {
        path.stem.lower(): path.name
        for path in level_directory.iterdir()
        if path.is_file() and IMAGE_RE.fullmatch(path.name)
    }


def _resolve_image_name(image: str | None, image_index: dict[str, str]) -> str | None:
    if not image:
        return None

    file_name = Path(image.strip().rstrip("/")).name
    if not IMAGE_RE.fullmatch(file_name):
        return None

    if not image_index:
        return file_name

    return image_index.get(Path(file_name).stem.lower())


def _parse_summaries(workbook: Any, level: str) -> dict[str, str]:
    worksheet = _find_summary_sheet(workbook, level)
    if worksheet is None:
        return {}

    summaries: dict[str, str] = {}

    for column in range(1, worksheet.max_column + 1):
        header = _cell_text(worksheet.cell(row=1, column=column).value)
        match = SUMMARY_HEADER_RE.search(header)
        if not match:
            continue

        summary = _cell_text(worksheet.cell(row=2, column=column).value)
        if summary:
            summaries[f"{int(match.group(1))}-{int(match.group(2))}"] = summary

    return summaries


def _split_image(content: str) -> tuple[str | None, str]:
    lines = content.splitlines()
    if not lines:
        return None, content

    first_line = lines[0].strip()
    if not IMAGE_RE.fullmatch(first_line):
        return None, content

    rest = lines[1:]
    while rest and not rest[0].strip():
        rest = rest[1:]

    return first_line, "\n".join(rest).strip()


def _parse_lesson_header(header: str) -> tuple[int, int, str] | None:
    match = LESSON_HEADER_RE.search(header)
    if not match:
        return None

    return int(match.group(1)), int(match.group(2)), match.group(3).strip()


def _lesson_group_columns(worksheet: Worksheet, start_column: int) -> dict[str, int]:
    columns = {
        "image": start_column,
        "content": start_column + 1,
        "script": start_column + 2,
        "vocabulary": start_column + 3,
    }

    for offset in range(4):
        column = start_column + offset
        header = _normalize(_cell_text(worksheet.cell(row=2, column=column).value))
        if not header:
            continue

        if IMAGE_TEXT_MARKER in header or "이미지" in header:
            columns["image"] = column
        elif "script" in header or _normalize(SCRIPT_TEXT_MARKER) in header:
            columns["script"] = column
        elif _normalize(VOCABULARY_TEXT_MARKER) in header:
            columns["vocabulary"] = column
        elif _normalize(LESSON_TEXT_MARKER) in header:
            columns["content"] = column

    return columns


def _cache_path(excel_path: Path) -> Path:
    return CACHE_DIR / f"{_normalize(excel_path.stem)}.json"


def _read_disk_cache(excel_path: Path, signature: int) -> dict[str, Any] | None:
    cache_path = _cache_path(excel_path)
    if not cache_path.exists():
        return None

    try:
        payload = json.loads(cache_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None

    if payload.get("cache_version") != CACHE_VERSION:
        return None

    if payload.get("source_mtime_ns") != signature:
        return None

    slides = payload.get("slides")
    if not isinstance(slides, dict):
        return None

    return slides


def _write_disk_cache(excel_path: Path, signature: int, slides: dict[str, Any]) -> None:
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    cache_path = _cache_path(excel_path)
    temp_path = cache_path.with_suffix(".tmp")
    payload = {
        "cache_version": CACHE_VERSION,
        "source": excel_path.name,
        "source_mtime_ns": signature,
        "slides": slides,
    }
    temp_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    temp_path.replace(cache_path)


def _parse_excel_slides(excel_path: Path, level: str) -> dict[str, Any]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    result: dict[str, Any] = {}

    try:
        worksheet = _find_lesson_sheet(workbook, level)
        if worksheet is None:
            return {}
        summaries = _parse_summaries(workbook, level)
        image_index = _lesson_image_index(level)

        for column in range(1, worksheet.max_column + 1):
            header = _cell_text(worksheet.cell(row=1, column=column).value)
            parsed_header = _parse_lesson_header(header)
            if not parsed_header:
                continue

            columns = _lesson_group_columns(worksheet, column)
            week, lesson, title = parsed_header
            lesson_key = f"{week}-{lesson}"
            slides: list[dict[str, Any]] = []

            for row in range(3, worksheet.max_row + 1):
                page_value = worksheet.cell(row=row, column=2).value
                slide_title = _cell_text(worksheet.cell(row=row, column=3).value)
                image = _cell_text(worksheet.cell(row=row, column=columns["image"]).value)
                content = _cell_text(worksheet.cell(row=row, column=columns["content"]).value)
                script = _cell_text(worksheet.cell(row=row, column=columns["script"]).value)
                vocabulary = _cell_text(worksheet.cell(row=row, column=columns["vocabulary"]).value)

                if not slide_title and not image and not content and not script and not vocabulary:
                    continue

                embedded_image, normalized_content = _split_image(content)
                resolved_image = _resolve_image_name(image or embedded_image, image_index)
                slides.append(
                    {
                        "page": _number(page_value, len(slides) + 1),
                        "title": slide_title,
                        "content": normalized_content,
                        "script": script,
                        "vocabulary": vocabulary,
                        "image": resolved_image,
                    }
                )

            result[lesson_key] = {
                "week": week,
                "lesson": lesson,
                "title": title,
                "summary": summaries.get(lesson_key, ""),
                "slides": slides,
            }
    finally:
        workbook.close()

    return result


def _load_level_slides(level: str) -> dict[str, Any]:
    cache_key = _normalize(level)
    excel_path = _find_excel_path(level)
    if excel_path is None:
        return {}

    signature = excel_path.stat().st_mtime_ns
    cached = _CACHE.get(cache_key)
    if cached and cached[0] == signature:
        return cached[1]

    disk_cached = _read_disk_cache(excel_path, signature)
    if disk_cached is not None:
        _CACHE[cache_key] = (signature, disk_cached)
        return disk_cached

    result = _parse_excel_slides(excel_path, level)
    if not result:
        return {}

    _CACHE[cache_key] = (signature, result)
    _write_disk_cache(excel_path, signature, result)
    return result


def warm_lecture_cache() -> None:
    for excel_path in _excel_candidates():
        signature = excel_path.stat().st_mtime_ns
        if _read_disk_cache(excel_path, signature) is not None:
            continue

        result = _parse_excel_slides(excel_path, excel_path.stem)
        if result:
            _write_disk_cache(excel_path, signature, result)


@router.get("/slides")
def list_level_slides(level: str = BEGINNER_1_TEXT) -> dict[str, Any]:
    return _load_level_slides(level)
